"""Excel report builder with formatting and conditional banding."""
from __future__ import annotations

import io
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .matcher import CandidateResult

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E78")
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="595959")
THIN = Side(border_style="thin", color="D9D9D9")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")


def _df_from_results(results: Iterable[CandidateResult]) -> pd.DataFrame:
    rows = [r.as_row() for r in results]
    df = pd.DataFrame(rows)
    if not df.empty:
        df.insert(0, "Rank", range(1, len(df) + 1))
    return df


def _autosize(ws, df: pd.DataFrame, header_row: int) -> None:
    for col_idx, col in enumerate(df.columns, start=1):
        header_len = len(str(col))
        sample = df[col].astype(str).head(50).map(len)
        max_len = max([header_len, *sample.tolist()]) if not sample.empty else header_len
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 60)
    ws.row_dimensions[header_row].height = 22


def _apply_score_bands(ws, df: pd.DataFrame, header_row: int) -> None:
    """Color the 'Overall Score' column green / yellow / red."""
    if "Overall Score" not in df.columns:
        return
    col_idx = list(df.columns).index("Overall Score") + 1
    letter = get_column_letter(col_idx)
    first = header_row + 1
    last = header_row + len(df)
    score_range = f"{letter}{first}:{letter}{last}"

    ws.conditional_formatting.add(
        score_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["70"], fill=GREEN_FILL),
    )
    ws.conditional_formatting.add(
        score_range,
        CellIsRule(operator="between", formula=["40", "69.999"], fill=YELLOW_FILL),
    )
    ws.conditional_formatting.add(
        score_range,
        CellIsRule(operator="lessThan", formula=["40"], fill=RED_FILL),
    )


def build_excel_report(
    results: list[CandidateResult],
    job_title: str = "Job Description",
    output_path: str | Path | None = None,
) -> bytes:
    """Render results to a styled Excel workbook.

    Args:
        results: Ranked candidate results.
        job_title: Used in the Summary sheet header.
        output_path: If provided, the workbook is also written to disk.

    Returns:
        Workbook bytes (suitable for download in Streamlit).
    """
    df = _df_from_results(results)
    wb = Workbook()

    # --- Sheet 1: Ranking ---
    ws = wb.active
    ws.title = "Ranking"

    ws["A1"] = f"Resume Screening Report — {job_title}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"{len(df)} candidates ranked by hybrid score "
        "(skill coverage 60% + semantic match 40%)."
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=max(len(df.columns), 1))
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=max(len(df.columns), 1))

    header_row = 4
    for col_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = CELL_BORDER

    for r_offset, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        for c_offset, value in enumerate(row.tolist(), start=1):
            cell = ws.cell(row=r_offset, column=c_offset, value=value)
            cell.border = CELL_BORDER
            if isinstance(value, str) and len(value) > 30:
                cell.alignment = WRAP
            elif isinstance(value, (int, float)):
                cell.alignment = CENTER

    if not df.empty:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=2)
        _autosize(ws, df, header_row)
        _apply_score_bands(ws, df, header_row)

    # --- Sheet 2: Methodology ---
    info = wb.create_sheet("Methodology")
    info.column_dimensions["A"].width = 28
    info.column_dimensions["B"].width = 80
    info["A1"] = "Resume Screener — Methodology"
    info["A1"].font = TITLE_FONT
    info.merge_cells("A1:B1")

    notes = [
        ("Overall Score", "Weighted blend: 60% Skill Coverage + 40% Semantic Match. Range 0–100."),
        ("Skill Coverage", "Share of JD skills (from a curated taxonomy) found in the resume."),
        ("Semantic Match", "TF-IDF cosine similarity of resume vs JD over 1–2 grams."),
        ("Matched Skills", "Skills present in BOTH the JD and the resume."),
        ("Missing Skills", "JD skills that did not appear in the resume."),
        ("Extra Skills", "Resume skills not requested by the JD — useful for adjacent roles."),
        ("Color Bands", "Green ≥ 70 | Yellow 40–69 | Red < 40."),
        ("Limitations", "Keyword-based; review top candidates manually before interview decisions."),
    ]
    for idx, (label, value) in enumerate(notes, start=3):
        info.cell(row=idx, column=1, value=label).font = Font(bold=True)
        cell = info.cell(row=idx, column=2, value=value)
        cell.alignment = WRAP
        info.row_dimensions[idx].height = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    payload = buffer.getvalue()

    if output_path:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "wb") as fh:
            fh.write(payload)

    return payload
